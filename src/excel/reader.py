from pathlib import Path
from datetime import datetime, date, time as dtime
import polars as pl
from openpyxl import load_workbook

def transform_excel_to_df(path: Path) -> pl.DataFrame:
    wb = None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        max_row, max_col = ws.max_row, ws.max_column

        used_cols = [False] * max_col
        rows = []

        iter_values = ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True)

        for r_idx, row in enumerate(iter_values, start=1):
            
            row_vals = list(row)

            if all(v is None or v == "" for v in row_vals):
                continue

            vals = []
            for j, v in enumerate(row_vals, start=1):
                if v is None or v == "":
                    vals.append(None)
                    continue

                used_cols[j - 1] = True

                if isinstance(v, (datetime, date, dtime)):
                    vals.append(v.isoformat())
                else:
                    s = str(v).strip()
                    vals.append(s if s != "" else None)
                    if s == "":
                        used_cols[j - 1] = False

            rows.append([r_idx, *vals])

        keep = [i + 1 for i, ok in enumerate(used_cols) if ok]
        cropped = []

        for r in rows:
            excel_row = r[0]
            data = r[1:]
            cropped.append([excel_row] + [data[i - 1] for i in keep])

        schema = ["EXCEL_ROW"] + [f"col_{i}" for i in keep]
        return pl.DataFrame(cropped, schema=schema, orient="row")
    
    finally:
        if wb is not None:
            wb.close()