from copy import copy
from openpyxl import load_workbook

from src.models.ticket_job import TicketJob
from src.models.excel_models import ExcelResult

from src.excel.transformers import split_web_creation_dt
from src.excel.validators import validate_excel


class ExcelWriter:
    def __init__(self, excel_result: ExcelResult):
        self.excel_path = excel_result.excel_path
        self.col_map = excel_result.col_map

        validate_excel(self.excel_path)

        self.wb = load_workbook(self.excel_path)
        self.wb.iso_dates = False
        self.ws = self.wb.worksheets[0]

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        self.close()

    def save(self):
        self.wb.save(self.excel_path)

    def save_as(self, output_path):
        self.wb.save(output_path)

    def close(self):
        try:
            self.wb.close()
        except Exception:
            pass

    # =========================
    # TICKET FLOW
    # =========================
    def add_datetime(self, job: TicketJob) -> None:
        if not job.creation_dt_text:
            return

        web_date, _ = split_web_creation_dt(job.creation_dt_text)
        self._write_if_empty(
            row=int(job.row_id),
            header="FECHA",
            value=web_date,
            number_format="dd-mm-yyyy",
        )

    def add_time(self, job: TicketJob) -> None:
        if not job.creation_dt_text:
            return

        _, web_time = split_web_creation_dt(job.creation_dt_text)
        self._write_if_empty(
            row=int(job.row_id),
            header="HORA",
            value=web_time,
            number_format="hh:mm:ss",
        )

    def add_ticket(self, job: TicketJob) -> None:
        if not job.ticket_id:
            return

        self._write_if_empty(
            row=int(job.row_id),
            header="TICKET",
            value=job.ticket_id,
        )
    

    # =========================
    # HELPERS
    # =========================
    def _write_if_empty(self, row: int, header: str, value, number_format: str | None = None):
        col = self.col_map[header]
        cell = self.ws.cell(row=row, column=col)

        if cell.value not in (None, "", "NONE"):
            return False

        self._change_style(row, col, cell)
        cell.value = value

        if number_format:
            cell.number_format = number_format
        return True

    def _change_style(self, row: int, col: int, cell) -> None:
        ref = self._find_reference_cell(row, col)
        if not ref:
            return

        cell.font = copy(ref.font)
        cell.fill = copy(ref.fill)
        cell.border = copy(ref.border)
        cell.alignment = copy(ref.alignment)
        cell.protection = copy(ref.protection)

    def _find_reference_cell(self, row: int, col: int):
        max_row = self.ws.max_row

        for r in range(row + 1, max_row + 1):
            ref = self.ws.cell(row=r, column=col)
            if ref.has_style:
                return ref

        return None